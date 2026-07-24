import os
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Define output directory
OUTPUT_DIR = r"c:\Users\namde\OneDrive\Desktop\forge\knowledge\data"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Common columns for all sheets
COMMON_COLS = [
    "id", "slug", "name", "canonical_name", "aliases", "category", "description",
    "developer", "organization", "license", "open_source", "sources",
    "schema_version", "data_version", "verified", "confidence_score", "created_at", "updated_at",
    "best_for", "not_recommended_for", "advantages", "limitations",
    "implementation_complexity", "learning_curve", "maintenance_effort"
]

# Specific columns mapping
SCHEMAS = {
    "llms.xlsx": [
        "model_type", "context_window", "supports_multimodal", "supports_function_calling",
        "supports_structured_output", "supports_multilingual", "reasoning_capability",
        "coding_capability", "latency", "throughput", "deployment_options", "pricing",
        "compatible_frameworks", "compatible_prompting_strategies", "compatible_agents", "requires_gpu"
    ],
    "embeddings.xlsx": [
        "embedding_dimension", "max_input_tokens", "supports_multilingual", "modalities",
        "latency", "memory_usage", "deployment_options", "compatible_vector_databases",
        "compatible_retrieval_strategies", "token_limit_sensitive"
    ],
    "vectordb.xlsx": [
        "features", "latency", "throughput", "scalability", "memory_usage", "deployment_options",
        "pricing", "compatible_embeddings", "compatible_frameworks", "compatible_retrieval_strategies",
        "requires_large_memory"
    ],
    "frameworks.xlsx": [
        "primary_paradigm", "programming_languages", "supports_rag", "supports_agents",
        "supports_tool_calling", "supports_streaming", "is_production_ready", "deployment_options",
        "compatible_llms", "compatible_vector_databases"
    ],
    "retrieval.xlsx": [
        "retrieval_type", "search_method", "supports_metadata_filtering", "supports_hybrid_search",
        "requires_reranker", "latency", "scalability", "compatible_vector_databases", "compatible_frameworks"
    ],
    "chunking.xlsx": [
        "chunking_type", "splitting_method", "preserves_semantics", "supports_overlap",
        "supports_hierarchical", "processing_speed", "token_efficiency", "compatible_embedding_models",
        "requires_document_structure"
    ],
    "rerankers.xlsx": [
        "reranker_type", "scoring_method", "is_llm_based", "supports_multilingual",
        "supports_long_documents", "latency", "inference_cost", "ranking_quality",
        "compatible_embedding_models", "compatible_retrieval_strategies", "requires_gpu"
    ],
    "prompting.xlsx": [
        "prompting_type", "is_reasoning_based", "supports_chain_of_thought", "supports_tool_calling",
        "token_efficiency", "reasoning_quality", "hallucination_resistance", "compatible_llms",
        "compatible_agents", "requires_reasoning_model"
    ],
    "agents.xlsx": [
        "agent_type", "is_multi_agent", "is_autonomous", "supports_memory", "supports_tool_calling",
        "supports_task_decomposition", "planning_capability", "execution_efficiency", "reliability",
        "compatible_llms", "compatible_frameworks", "requires_high_reasoning_model"
    ],
    "evaluation.xlsx": [
        "evaluation_type", "is_automated", "is_llm_as_judge", "evaluates_answer_correctness",
        "evaluates_hallucination", "evaluates_context_precision", "evaluates_retrieval_precision",
        "evaluation_speed", "computational_cost", "compatible_frameworks", "requires_ground_truth"
    ]
}

# Validation rule setup
bool_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
bool_dv.error = 'Value must be TRUE or FALSE'

score_dv = DataValidation(type="whole", operator="between", formula1=1, formula2=10, allow_blank=True)
score_dv.error = 'Value must be between 1 and 10'

conf_dv = DataValidation(type="decimal", operator="between", formula1=0.0, formula2=1.0, allow_blank=True)
conf_dv.error = 'Value must be a decimal between 0.0 and 1.0'

for filename, spec_cols in SCHEMAS.items():
    wb = openpyxl.Workbook()
    # Ensure sheet is named Data
    ws = wb.active
    ws.title = "Data"
    
    # Combine headers
    headers = COMMON_COLS + spec_cols
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        # auto fit rough width
        ws.column_dimensions[get_column_letter(col_idx)].width = len(header) + 5

    # Enable filter and freeze first row
    max_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{max_col_letter}1000"
    ws.freeze_panes = "A2"
    
    # Apply validations
    ws.add_data_validation(bool_dv)
    ws.add_data_validation(score_dv)
    ws.add_data_validation(conf_dv)

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}1000"
        
        # Apply boolean validation
        if header.startswith(("supports_", "is_", "requires_", "evaluates_", "open_source", "verified")):
            bool_dv.add(cell_range)
        
        # Apply 1-10 score validation
        elif header in [
            "implementation_complexity", "learning_curve", "maintenance_effort",
            "reasoning_capability", "coding_capability", "latency", "throughput", "scalability",
            "memory_usage", "processing_speed", "token_efficiency", "inference_cost", 
            "ranking_quality", "reasoning_quality", "hallucination_resistance", 
            "planning_capability", "execution_efficiency", "reliability", "evaluation_speed", 
            "computational_cost"
        ]:
            score_dv.add(cell_range)
        
        # Apply confidence score validation
        elif header == "confidence_score":
            conf_dv.add(cell_range)

    # Save workbook
    wb.save(os.path.join(OUTPUT_DIR, filename))
    print(f"Generated {filename}")

print("All Excel templates generated successfully.")
